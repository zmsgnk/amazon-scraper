#!/usr/bin/env python
# -*- coding: utf-8 -*-

from bs4 import BeautifulSoup
import urllib2
from openpyxl import Workbook
from datetime import datetime
import time
import sys


def scrape_bestseller(url): 
	html = urllib2.urlopen(url).read().decode('shift-jis')
	soup = BeautifulSoup(html)	

	rank = 1
	result = []
	for item in soup.findAll('div', {'class': 'zg_item_normal'}): 
		tmp = []

		## ranking
		tmp.append(rank)
		rank += 1

		## brand
		try:
			brand = item.find('div', {'class': 'zg_byline'}).string
			brand = brand.replace('\n', '')
		except AttributeError:
			brand = "NA"
		tmp.append(brand)

		## title
		try:
			zg_title = item.find('div', {'class': 'zg_title'})
			title = zg_title.string.replace('\n', '')
		except AttributeError: 
			title = "NA"
		tmp.append(title)

		## link
		try:
			link = zg_title.find('a').get('href')
			link = link.replace('\n', '')
		except AttributeError:
			link = "NA"
		tmp.append(link)

		## list price
		try:
			list_price = item.find('span', {'class': 'listprice'}).string
			# list_price = prettify_price(list_price)
		except AttributeError:
			list_price = 'NA'
		tmp.append(list_price)

		## price
		try:
			price = item.find('span', {'class': 'price'}).find('b').string
  		    # price = prettify_price(price)
  		except AttributeError:
  			price = "NA"
		tmp.append(price)

		result.append(tmp)
		
	return result


def prettify_price(price):
	price = price.replace(u'\uffe5 ', '')
	price = price.replace(',', '')
	price = int(price)
	return(price)

		
if __name__ == '__main__':
	
	categories = ["electronics", "food-beverage", "hpc", "automotive", 
	              "beauty", "shoes", "sports", "computers", "baby", 
	              "appliances", "apparel", "watch"]

	data = {}
	for category in categories:
		tmp = [["ランキング", "ブランド", "商品名", "URL", "参考価格", "価格"]]
		for page_num in xrange(5):
			url = "http://www.amazon.co.jp/gp/bestsellers/%s/#%s" % (category, str(page_num+1))
			result = scrape_bestseller(url)
			tmp.extend(result)
			time.sleep(1)
		data[category] = tmp

	## write xlsx
	wb = Workbook()
	ws = wb.worksheets[0]
	for category in categories:
		ws.title = category
		for i, item_list in enumerate(data[category]):
			for j, item in enumerate(item_list):
				if j == 2 and i != 0:
					ws.cell(row=i+1, column=j+1).hyperlink = item
				else:
					ws.cell(row=i+1, column=j+1).value = item
		ws = wb.create_sheet()

	today = datetime.today()
	path_to_file = sys.argv[1]
	fname = path_to_file + "amazon-bestseller_%s.xlsx" % today.strftime("%Y%m%d")

	wb.save(filename=fname)